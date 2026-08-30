#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importa los microciclos reales (Excel) y genera data.js para la app de carga GPS.

Fuente: ~/Desktop/AT BALEARES 26-27/GPS/
  - Microciclo_Tipo.xlsx  (hojas MICROCICLOS = coeficientes, REF_PARTIDO)
  - Microciclos/Microciclo 1..6.xlsx  (hojas Sxx_GPS, PTx_GPS, CARGAS_OBJETIVO|Acumulado, CARGA_AC)

Uso:  python3 import_data.py
Salida:  data.js  ->  window.GPS_DATA_ALL = { meta, refPartido, coeficientes, microciclos, M1..M6 }

La app SOLO muestra lo que hay en el Excel (obj / real / dif / ACWR / medias ya calculados).
Lo único que calcula este script: el 'real' acumulado solo-sesiones (el Excel mezcla el
partido) y la serie diaria de Player Load de 28 días para la gráfica de Carga A:C.
"""
import json, re, glob, os, datetime as dt
import openpyxl

GPS_DIR = os.path.expanduser("~/Desktop/AT BALEARES 26-27/GPS")
MICRO_DIR = os.path.join(GPS_DIR, "Microciclos")
TIPO_XLSX = os.path.join(GPS_DIR, "Microciclo_Tipo.xlsx")
OUT = os.path.join(os.path.dirname(__file__), "data.js")

METRICS = ["distancia", "hmld", "hsr", "sprint", "acc", "dec"]
MES = {1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
       7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC"}


# ---------------------------------------------------------------- utilidades
def num(v):
    if v is None or v == "" or v == "·":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace("\xa0", "").replace(",", ".")
    try:
        f = float(s)
        return int(f) if f == int(f) else round(f, 2)
    except ValueError:
        return None


def find_date(text):
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text or "")
    if not m:
        return None
    d, mo, y = map(int, m.groups())
    return dt.date(y, mo, d)


def iso(d):
    return d.isoformat() if d else None


def hms_to_s(t):
    if not t:
        return None
    parts = str(t).strip().split(":")
    try:
        parts = [int(float(x)) for x in parts]
    except ValueError:
        return None
    if len(parts) == 3:
        return parts[0] * 3600 + parts[1] * 60 + parts[2]
    if len(parts) == 2:
        return parts[0] * 60 + parts[1]
    return None


def s_to_hms(s):
    s = int(round(s))
    return "%d:%02d:%02d" % (s // 3600, (s % 3600) // 60, s % 60)


def team_extras(team, players):
    """Rellena la media de velMax / playerLoad y la duración del equipo a partir de los
    jugadores participantes (el Excel no trae estas medias)."""
    parts = [p for p in players if p.get("estado") != "na"]

    def mean(key):
        vals = [p.get(key) for p in parts if p.get(key) is not None]
        return sum(vals) / len(vals) if vals else None

    vmax = mean("velMax")
    pl = mean("playerLoad")
    durs = [hms_to_s(p.get("duracion")) for p in parts]
    durs = [d for d in durs if d]
    if team is None and (vmax is not None or pl is not None):
        team = {}
    if team is not None:
        if team.get("velMax") is None and vmax is not None:
            team["velMax"] = round(vmax, 1)
        if team.get("playerLoad") is None and pl is not None:
            team["playerLoad"] = round(pl)
        if team.get("duracion") is None and durs:
            team["duracion"] = s_to_hms(max(durs))
    return team


def header_row(ws, key="Dorsal"):
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and str(row[0]).strip() == key:
            return i, row
    return None, None


def player_rows(ws, hdr_idx):
    """Filas de jugador (col A = dorsal numérico) y la fila MEDIA EQUIPO si existe."""
    rows, media = [], None
    for row in ws.iter_rows(min_row=hdr_idx + 1, values_only=True):
        a, b = row[0], (row[1] if len(row) > 1 else None)
        if a is not None and str(a).strip().isdigit():
            rows.append(row)
        elif b and "MEDIA" in str(b).upper():
            media = row
        elif media is None and rows and not a and not b and any(
                isinstance(c, (int, float)) for c in row[3:]):
            media = row  # fila resumen sin etiqueta (justo tras los jugadores)
    return rows, media


def role_from_title(a3, a2):
    """Extrae el rol del día (MD-3, +1 Suplentes, Mixto, Other...) del título/nota."""
    t = (a3 or "")
    parts = [p.strip() for p in t.split("·")]
    if len(parts) >= 3:
        r = parts[2]
    else:
        r = ""
    r = r.strip()
    low = (r + " " + (a2 or "")).lower()
    if not r or r.upper().startswith("TIPO"):
        if "sesión \"other\"" in low or 'sesion "other"' in low or "sesiones \"other\"" in low:
            r = "Other"
        elif "mixto" in low:
            r = "Mixto"
        else:
            r = ""
    # normaliza
    rr = r.upper()
    if "SUPLENT" in rr:
        return "+1 Supl."
    if "TITULAR" in rr:
        return "+1 Tit."
    if rr.startswith("MD-") or re.match(r"-\d", rr):
        return rr.replace("MD-", "MD-") if rr.startswith("MD-") else "MD" + rr
    if "MIXTO" in rr:
        return "Mixto"
    if "OTHER" in rr:
        return "Other"
    return r or "—"


# ---------------------------------------------------------------- REF_PARTIDO + coeficientes
def load_tipo():
    wb = openpyxl.load_workbook(TIPO_XLSX, data_only=True)

    # --- coeficientes ---
    ws = wb["MICROCICLOS"]
    rows = list(ws.iter_rows(values_only=True))
    coef, cur = {}, None
    day_keys = ["+1 Suplentes", "+1 Titulares", "MD-4", "MD-3", "MD-2", "MD-1", "Acumulado Semana"]
    for row in rows:
        c0 = str(row[0] or "").strip()
        mt = re.match(r"MICROCICLO TIPO ([ABC])", c0)
        if mt:
            cur = mt.group(1)
            coef[cur] = {}
            continue
        if cur and c0.upper() in ("DISTANCIA", "HMLD", "HSR", "SPRINT", "ACC", "DEC"):
            mkey = {"SPRINT": "sprint"}.get(c0.upper(), c0.lower())
            for j, dk in enumerate(day_keys, start=1):
                v = num(row[j]) if j < len(row) else None
                if v is not None:
                    coef[cur].setdefault(dk, {})[mkey] = v
    # renombra días a claves cortas
    ren = {"+1 Suplentes": "+1S", "+1 Titulares": "+1T", "MD-4": "MD-4",
           "MD-3": "MD-3", "MD-2": "MD-2", "MD-1": "MD-1", "Acumulado Semana": "ACUM"}
    coef = {t: {ren[d]: mm for d, mm in days.items()} for t, days in coef.items()}

    # --- REF_PARTIDO ---
    ws = wb["REF_PARTIDO"]
    hi, hdr = header_row(ws)
    ref = {}
    notas = []
    for row in ws.iter_rows(min_row=hi + 1, values_only=True):
        a = row[0]
        if a is not None and str(a).strip().isdigit():
            ref[int(a)] = dict(
                dorsal=int(a), jugador=titlecase(row[1]), grupo=(row[2] or "").strip(),
                distancia=num(row[3]), hmld=num(row[4]), hsr=num(row[5]),
                sprint=num(row[6]), acc=num(row[7]), dec=num(row[8]))
        elif row[0] and not any(row[1:]):
            notas.append(str(row[0]))
    return coef, ref, notas


def titlecase(name):
    if not name:
        return name
    s = str(name).strip()
    # "MONTCHEU, F." -> "Montcheu, F."
    parts = s.split(",")
    ape = parts[0].strip().title()
    ini = parts[1].strip().upper() if len(parts) > 1 else ""
    return f"{ape}, {ini}" if ini else ape


# ---------------------------------------------------------------- sesiones / partidos
def parse_session(ws):
    a1 = ws.cell(1, 1).value or ""
    a2 = ws.cell(2, 1).value or ""
    a3 = ws.cell(3, 1).value or ""
    date = find_date(a1) or find_date(a2)
    tipo = None
    mt = re.search(r"TIPO ([ABC])", a3 + " " + a2)
    if mt:
        tipo = mt.group(1)
    role = role_from_title(a3, a2)
    hi, hdr = header_row(ws)
    prows, media = player_rows(ws, hi)

    def cell(row, base):  # base = índice de la col Obj
        o, r, d = num(row[base]), num(row[base + 1]), num(row[base + 2])
        if o is None and r is None and d is None:
            return None
        return dict(obj=o, real=r, dif=d)

    players = []
    for row in prows:
        rec = dict(dorsal=int(row[0]), jugador=titlecase(row[1]), grupo=(row[2] or "").strip())
        allnull = True
        for k, base in zip(METRICS, range(3, 3 + 18, 3)):
            c = cell(row, base)
            rec[k] = c
            if c and (c["real"] is not None or c["obj"] is not None):
                allnull = allnull and (c["real"] is None)
        rec["velMax"] = num(row[21])
        rec["playerLoad"] = num(row[22])
        rec["duracion"] = (str(row[23]).strip() if row[23] not in (None, "·", "") else None)
        # estado: sin ningún real ni duración -> no participó
        has_real = any(rec[k] and rec[k]["real"] is not None for k in METRICS) or rec["playerLoad"] is not None
        if not has_real:
            rec["estado"] = "na"
        players.append(rec)

    team = None
    if media:
        team = {}
        for k, base in zip(METRICS, range(3, 3 + 18, 3)):
            o, r, d = num(media[base]), num(media[base + 1]), num(media[base + 2])
            team[k] = dict(obj=o, real=r, dif=d)
        team["velMax"] = num(media[21])
        team["playerLoad"] = num(media[22])
    team = team_extras(team, players)

    return dict(date=iso(date), role=role, tipo=tipo, nota=a2, titulo=a1,
                players=players, teamAvg=team)


def parse_match(ws):
    a1 = ws.cell(1, 1).value or ""
    a2 = ws.cell(2, 1).value or ""
    date = find_date(a1) or find_date(a2)
    mr = re.search(r"vs ([^(·]+?)\s*[\(·]", a1)
    rival = mr.group(1).strip() if mr else ""
    hi, hdr = header_row(ws)
    prows, media = player_rows(ws, hi)
    players = []
    for row in prows:
        rec = dict(dorsal=int(row[0]), jugador=titlecase(row[1]), grupo=(row[2] or "").strip())
        vals = [num(row[3 + i]) for i in range(6)]
        for k, v in zip(METRICS, vals):
            rec[k] = dict(obj=None, real=v, dif=None) if v is not None else dict(obj=None, real=None, dif=None)
        rec["velMax"] = num(row[9])
        rec["playerLoad"] = num(row[10])
        rec["duracion"] = (str(row[11]).strip() if row[11] not in (None, "·", "") else None)
        if all(v is None for v in vals) and rec["playerLoad"] is None:
            rec["estado"] = "na"
        players.append(rec)
    team = None
    if media:
        team = {}
        for k, i in zip(METRICS, range(3, 9)):
            v = num(media[i])
            team[k] = dict(obj=None, real=v, dif=None)
        team["velMax"] = num(media[9])
        team["playerLoad"] = num(media[10])
    team = team_extras(team, players)
    return dict(date=iso(date), role="Partido", rival=rival, nota=a2, titulo=a1,
               players=players, teamAvg=team)


def parse_acumulado(ws):
    hi, hdr = header_row(ws)
    prows, media = player_rows(ws, hi)
    a3 = ws.cell(3, 1).value or ""
    a4 = ws.cell(4, 1).value or ""

    def rowrec(row):
        rec = dict(dorsal=int(row[0]), jugador=titlecase(row[1]), grupo=(row[2] or "").strip())
        for k, base in zip(METRICS, range(3, 3 + 18, 3)):
            rec[k] = dict(obj=num(row[base]), real=num(row[base + 1]), dif=num(row[base + 2]))
        return rec

    players = [rowrec(r) for r in prows]
    team = None
    if media:
        team = {}
        for k, base in zip(METRICS, range(3, 3 + 18, 3)):
            team[k] = dict(obj=num(media[base]), real=num(media[base + 1]), dif=num(media[base + 2]))
    return dict(players=players, teamAvg=team, semana=a3, nota=a4)


def parse_carga_ac(ws):
    hi, hdr = header_row(ws)
    # columnas dinámicas: ACWR (idx3), luego PL Sxx..., aguda, cronica
    pl_cols = []
    for j in range(4, len(hdr)):
        h = str(hdr[j] or "")
        m = re.search(r"PL\s+(S\d+\w*|PT\d+)", h)
        if m:
            pl_cols.append((j, m.group(1)))
    aguda_col = len(hdr) - 2
    cron_col = len(hdr) - 1
    prows, media = player_rows(ws, hi)
    players = []
    for row in prows:
        rec = dict(dorsal=int(row[0]), jugador=titlecase(row[1]), grupo=(row[2] or "").strip(),
                   acwr=num(row[3]),
                   cargaAguda=num(row[aguda_col]), cargaCronica=num(row[cron_col]))
        for j, key in pl_cols:
            rec["pl" + key] = num(row[j])
        players.append(rec)
    team = None
    if media:
        team = dict(acwr=num(media[3]),
                    cargaAguda=num(media[aguda_col]), cargaCronica=num(media[cron_col]))
    a2 = ws.cell(2, 1).value or ""
    a3 = ws.cell(3, 1).value or ""
    calc = find_date(a2)
    if not calc:
        m = re.search(r"[Ff]echa de c[aá]lculo[^(]*\((\d{1,2})/(\d{1,2})", a2)
        if m:
            calc = dt.date(2026, int(m.group(2)), int(m.group(1)))
    return dict(players=players, teamAvg=team, nota=(a2 + " " + a3).strip(),
                calcISO=iso(calc), plCols=[k for _, k in pl_cols])


# ---------------------------------------------------------------- microciclos
def week_label(a3, n):
    m = re.search(r"SEMANA:\s*([^·]+)", a3 or "")
    if m:
        return m.group(1).strip()
    m = re.search(r"MICROCICLO\s*\d+\s*\(([^)]+)\)", a3 or "")
    return (m.group(1).strip() if m else f"Microciclo {n}")


def load_microcycle(path):
    n = int(re.search(r"Microciclo (\d+)", path).group(1))
    wb = openpyxl.load_workbook(path, data_only=True)
    sesiones, partidos, orden = {}, {}, []
    for name in wb.sheetnames:
        if name.endswith("_GPS") and name.startswith("PT"):
            key = name[:-4]
            partidos[key] = parse_match(wb[name])
            orden.append(("partido", key, partidos[key]["date"]))
        elif name.endswith("_GPS"):
            key = name[:-4]
            # S6a/S6b -> se mantienen separadas (así están en el Excel)
            sesiones[key] = parse_session(wb[name])
            orden.append(("sesion", key, sesiones[key]["date"]))
    orden.sort(key=lambda t: (t[2] or "9999"))

    acu_name = "Acumulado" if "Acumulado" in wb.sheetnames else "CARGAS_OBJETIVO"
    acu = parse_acumulado(wb[acu_name])
    cac = parse_carga_ac(wb["CARGA_AC"])

    tipo = None
    for _, k, _ in orden:
        s = sesiones.get(k)
        if s and s["tipo"]:
            tipo = s["tipo"]
            break

    ses_keys = [k for kind, k, _ in orden if kind == "sesion"]

    def any_load(src):
        return src is not None and any(p.get("playerLoad") is not None for p in src["players"])
    done_dates = [dat for kind, k, dat in orden
                  if dat and any_load(sesiones.get(k) or partidos.get(k))]

    a3 = acu["semana"]
    # el día real del último dato manda sobre la nota "fecha de cálculo" (que a veces viene desfasada)
    calc = (max(done_dates) if done_dates
            else (cac["calcISO"] or (orden[-1][2] if orden else None)))

    # el "activo" lo decide main() (siempre el microciclo más reciente); aquí solo la meta base
    meta = dict(n=n, titulo=f"Microciclo {n}", tipo=tipo,
                semana=week_label(a3, n), temporada="2026-27",
                calculoISO=calc,
                calculoFecha=(dt.date.fromisoformat(calc).strftime("%d/%m") if calc else "—"),
                estado="cerrado")

    # cargasObjetivo (solo sesiones): obj = del Excel (ya es solo-sesiones), real = Σ sesiones
    cargas_obj = build_sessions_only(sesiones, ses_keys, acu)

    return n, dict(
        meta=meta,
        orden=[dict(tipo=kind, key=k) for kind, k, _ in orden],
        sesiones=sesiones, partidos=partidos,
        cargasObjetivo=cargas_obj,
        cargasSemana=dict(players=acu["players"], teamAvg=acu["teamAvg"],
                          nota="Acumulado de toda la semana: sesiones de entrenamiento + partido(s). " + acu["nota"]),
        cargaAC=dict(players=cac["players"], teamAvg=cac["teamAvg"], nota=cac["nota"]),
    ), cac


def build_sessions_only(sesiones, ses_keys, acu):
    obj_by = {p["dorsal"]: p for p in acu["players"]}
    players = []
    for dor, ap in obj_by.items():
        rec = dict(dorsal=dor, jugador=ap["jugador"], grupo=ap["grupo"])
        for k in METRICS:
            o = ap[k]["obj"]
            real = 0
            has = False
            for sk in ses_keys:
                pr = next((x for x in sesiones[sk]["players"] if x["dorsal"] == dor), None)
                if pr and pr[k] and pr[k]["real"] is not None:
                    real += pr[k]["real"]
                    has = True
            rec[k] = dict(obj=o, real=(round(real, 2) if has else None),
                          dif=(round(real - o, 2) if (has and o is not None) else None))
        players.append(rec)
    # team avg
    team = {}
    for k in METRICS:
        o = acu["teamAvg"][k]["obj"] if acu["teamAvg"] else None
        reals = [p[k]["real"] for p in players if p[k]["real"] is not None and p["grupo"]]
        r = round(sum(reals) / len(reals)) if reals else None
        team[k] = dict(obj=o, real=r, dif=(None if (o is None or r is None) else round(r - o)))
    return dict(players=players, teamAvg=team,
                nota="Objetivo acumulado de las sesiones de entrenamiento de la semana (sin partido).")


# ---------------------------------------------------------------- serie diaria de Player Load
def build_series(all_days, micro_data, cac_by_n):
    """all_days: dict iso_date -> {dorsal: pl}. Rellena cargaAC[i].serie de CADA microciclo
    con fecha de cálculo: ventana de 28 días que acaba en su fecha de cálculo. El activo la usa
    en la pestaña Carga A:C; los cerrados, en Historial."""
    for n, m in micro_data.items():
        calc = m["meta"]["calculoISO"]
        if not calc:
            continue
        asof = dt.date.fromisoformat(calc)
        dias = [(asof - dt.timedelta(days=k)) for k in range(27, -1, -1)]
        m["cargaAC"]["serieDias"] = [d.isoformat() for d in dias]
        # etiqueta de sesión por día (según ESTE microciclo)
        seslabel = {}
        for kind, k, _ in [(o["tipo"], o["key"], None) for o in m["orden"]]:
            src = m["sesiones"].get(k) or m["partidos"].get(k)
            if src and src["date"]:
                seslabel[src["date"]] = k
        for p in m["cargaAC"]["players"]:
            dor = p["dorsal"]
            pl = []
            for d in dias:
                pl.append(all_days.get(d.isoformat(), {}).get(dor, 0) or 0)
            aguda, cron, acwr, ses = [], [], [], []
            for i, d in enumerate(dias):
                w = sum(all_days.get((d - dt.timedelta(days=k)).isoformat(), {}).get(dor, 0) or 0 for k in range(7))
                mo = sum(all_days.get((d - dt.timedelta(days=k)).isoformat(), {}).get(dor, 0) or 0 for k in range(28))
                a, c = round(w / 7), round(mo / 28)
                aguda.append(a)
                cron.append(c)
                acwr.append(round(a / c, 2) if c else 0.0)
                ses.append(seslabel.get(d.isoformat(), "") if pl[i] else "")
            # el último punto lo fija el Excel (los microciclos anteriores pueden estar
            # incompletos en nuestros datos y desviar la crónica reconstruida)
            if p.get("cargaAguda") is not None:
                aguda[-1] = round(p["cargaAguda"])
            if p.get("cargaCronica") is not None:
                cron[-1] = round(p["cargaCronica"])
            if p.get("acwr") is not None:
                acwr[-1] = p["acwr"]
            p["serie"] = dict(pl=pl, aguda=aguda, cronica=cron, acwr=acwr, ses=ses)

        # serie de la MEDIA DEL EQUIPO (media diaria de los jugadores de campo con datos)
        field = [p for p in m["cargaAC"]["players"] if p["grupo"]]
        def col_mean(key, i):
            vals = [p["serie"][key][i] for p in field if p["serie"][key][i]]
            return round(sum(vals) / len(vals)) if vals else 0
        tpl = [col_mean("pl", i) for i in range(28)]
        tag_ = [col_mean("aguda", i) for i in range(28)]
        tcr = [col_mean("cronica", i) for i in range(28)]
        tac = [round(tag_[i] / tcr[i], 2) if tcr[i] else 0.0 for i in range(28)]
        tse = [seslabel.get(d.isoformat(), "") if tpl[i] else "" for i, d in enumerate(dias)]
        ta = m["cargaAC"].get("teamAvg") or {}
        if ta.get("cargaAguda") is not None:
            tag_[-1] = round(ta["cargaAguda"])
        if ta.get("cargaCronica") is not None:
            tcr[-1] = round(ta["cargaCronica"])
        if ta.get("acwr") is not None:
            tac[-1] = ta["acwr"]
        m["cargaAC"]["serieTeam"] = dict(pl=tpl, aguda=tag_, cronica=tcr, acwr=tac, ses=tse)


# ---------------------------------------------------------------- main
def main():
    coef, ref, ref_notas = load_tipo()

    paths = sorted(glob.glob(os.path.join(MICRO_DIR, "Microciclo *.xlsx")),
                   key=lambda p: int(re.search(r"(\d+)", os.path.basename(p)).group(1)))
    micro_data = {}
    cac_by_n = {}
    all_days = {}   # iso -> {dorsal: pl}
    velmax_match = {}
    velmax_sess = {}
    partidos_jugados = {}

    for path in paths:
        n, m, cac = load_microcycle(path)
        micro_data[n] = m
        cac_by_n[n] = cac
        # recolecta PL diario + velmax + partidos
        for k, s in list(m["sesiones"].items()):
            for p in s["players"]:
                if s["date"] and p.get("playerLoad") is not None:
                    all_days.setdefault(s["date"], {})[p["dorsal"]] = p["playerLoad"]
                if p.get("velMax") is not None:
                    velmax_sess[p["dorsal"]] = max(velmax_sess.get(p["dorsal"], 0), p["velMax"])
        for k, s in list(m["partidos"].items()):
            valido_ref = k not in ("PT4",)   # PT4 anulado como referencia de partido
            for p in s["players"]:
                if s["date"] and p.get("playerLoad") is not None:
                    all_days.setdefault(s["date"], {})[p["dorsal"]] = p["playerLoad"]
                if valido_ref and p.get("velMax") is not None:
                    velmax_match[p["dorsal"]] = max(velmax_match.get(p["dorsal"], 0), p["velMax"])
                if valido_ref and p.get(METRICS[0]) and p[METRICS[0]]["real"] is not None:
                    partidos_jugados[p["dorsal"]] = partidos_jugados.get(p["dorsal"], 0) + 1

    build_series(all_days, micro_data, cac_by_n)

    # nº de partidos que hay DETRÁS de la tabla REF_PARTIDO. Según las notas de la hoja,
    # la referencia vigente es la media de PT1-PT3, PT5, PT6, PT7 y PT8 (PT4 anulado;
    # días "Modified" y la sustitución de GPS de Hernández fuera).
    REF_MATCHES = {"PT1", "PT2", "PT3", "PT5", "PT6", "PT7", "PT8"}
    NO_REF = {(17, "PT2"), (24, "PT2"), (24, "PT3"), (23, "PT3"), (14, "PT3")}
    partidos_ref = {}
    for n, m in sorted(micro_data.items()):
        for k, s in m["partidos"].items():
            if k not in REF_MATCHES:
                continue
            for p in s["players"]:
                if (p["dorsal"], k) in NO_REF:
                    continue
                if p.get(METRICS[0]) and p[METRICS[0]]["real"] is not None:
                    partidos_ref[p["dorsal"]] = partidos_ref.get(p["dorsal"], 0) + 1

    # REF_PARTIDO: añade velMax (mejor de partido, si no de sesión) y nº de partidos
    ref_players = []
    for dor, r in sorted(ref.items()):
        r = dict(r)
        r["velMax"] = velmax_match.get(dor) or velmax_sess.get(dor)
        r["partidos"] = partidos_ref.get(dor, partidos_jugados.get(dor, 0))
        ref_players.append(r)
    field = ref_players
    tavg = {}
    for k in METRICS + ["velMax"]:
        vals = [p[k] for p in field if p.get(k) is not None]
        tavg[k] = round(sum(vals) / len(vals), 1 if k == "velMax" else 0)
        if k != "velMax":
            tavg[k] = int(tavg[k])

    # el microciclo más reciente es el "activo"
    n_activo = max(micro_data)
    micro_data[n_activo]["meta"]["estado"] = "activo"

    micro_keys = [f"M{n}" for n in sorted(micro_data, reverse=True)]
    DATA = {
        "meta": {
            "temporada": "2026-27",
            "club": "AT BALEARES",
            "generado": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "fuente": "AT BALEARES 26-27/GPS (Microciclo 1-%d)" % max(micro_data),
        },
        "refPartido": {
            "nota": ("REF_PARTIDO = media de los partidos válidos de cada jugador, cada uno estimado a 95' "
                     "con fórmula de fatiga (PT1-PT3, PT5, PT6, PT7 y PT8; PT4 anulado; días 'Modified' fuera). "
                     "La media del equipo se calcula con toda la plantilla de campo. "
                     "Vel. máx tomada del mejor registro de partido."),
            "players": ref_players,
            "teamAvg": tavg,
            "notas": ref_notas,
        },
        "coeficientes": {"nota": "Coeficiente de carga por Tipo de microciclo y día (hoja MICROCICLOS).",
                         "tipos": coef},
        "microciclos": micro_keys,
    }
    for n, m in micro_data.items():
        DATA[f"M{n}"] = m

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("/* Datos reales de carga GPS — AT BALEARES 26-27. Generado por import_data.py. */\n")
        f.write("/* NO editar a mano: se regenera desde los Excel de ~/Desktop/AT BALEARES 26-27/GPS/. */\n")
        f.write("window.GPS_DATA_ALL = ")
        json.dump(DATA, f, ensure_ascii=False, indent=1)
        f.write(";\n")

    print("data.js generado.")
    for n in sorted(micro_data):
        m = micro_data[n]
        print(f"  M{n} [{m['meta']['estado']}] {m['meta']['semana']} · "
              f"{len(m['sesiones'])} sesiones + {len(m['partidos'])} partidos · calc {m['meta']['calculoFecha']}")
    print("  jugadores REF_PARTIDO:", len(ref_players))


if __name__ == "__main__":
    main()
